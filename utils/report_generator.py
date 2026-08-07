import pandas as pd
from  io import BytesIO
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def generate_csv_report(df):
    return df.to_csv(index=False).encode("utf-8-sig")


def generate_excel_report(df, kpis):
    output = BytesIO()

    product_performance = (
        df.groupby("Product")[["Revenue", "Profit"]]
        .sum()
        .reset_index()
        .sort_values("Revenue", ascending=False)
    )

    regional_analysis = (
        df.groupby("Region")[["Revenue", "Profit"]]
        .sum()
        .reset_index()
        .sort_values("Revenue", ascending=False)
    )

    kpi_df = pd.DataFrame(
        {
            "Metric": [
                "Total Revenue",
                "Total Profit",
                "Total Orders",
                "Average Order Value",
                "Profit Margin"
            ],
            "Value": [
                kpis["total_revenue"],
                kpis["total_profit"],
                kpis["total_orders"],
                kpis["average_order_value"],
                kpis["profit_margin"]
            ]
        }
    )

    executive_summary = pd.DataFrame(
    {
        "Metric": [
            "Total Revenue",
            "Total Profit",
            "Total Orders",
            "Average Order Value",
            "Profit Margin"
        ],
        "Value": [
            kpis["total_revenue"],
            kpis["total_profit"],
            kpis["total_orders"],
            kpis["average_order_value"],
            kpis["profit_margin"]
        ]
    }
)

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        executive_summary.to_excel(
            writer,
            sheet_name="Executive Summary",
            index=False
        )

        kpi_df.to_excel(
            writer,
            sheet_name="KPIs",
            index=False
        )

        product_performance.to_excel(
            writer,
            sheet_name="Product Performance",
            index=False
        )

        regional_analysis.to_excel(
            writer,
            sheet_name="Regional Analysis",
            index=False
        )

        df.to_excel(
            writer,
            sheet_name="Filtered Data",
            index=False
        )

        for sheet_name in writer.book.sheetnames:
            worksheet = writer.book[sheet_name]

            for cell in worksheet[1]:
                cell.font = Font(bold=True)

            for column_cells in worksheet.columns:
                max_length = 0

                column_letter = get_column_letter(
                    column_cells[0].column
                )

                for cell in column_cells:
                    try:
                        if cell.value is not None:
                            max_length = max(
                                max_length,
                                len(str(cell.value))
                            )
                    except TypeError:
                        pass

                worksheet.column_dimensions[column_letter].width = (
                    min(max_length + 2, 30)
                )

        product_sheet = writer.book["Product Performance"]

        for row in range(2, product_sheet.max_row + 1):
            product_sheet[f"B{row}"].number_format = '$#,##0.00'
            product_sheet[f"C{row}"].number_format = '$#,##0.00'

        regional_sheet = writer.book["Regional Analysis"]

        for row in range(2, regional_sheet.max_row + 1):
            regional_sheet[f"B{row}"].number_format = '$#,##0.00'
            regional_sheet[f"C{row}"].number_format = '$#,##0.00'

        kpi_sheet = writer.book["KPIs"]

        for row in range(2, kpi_sheet.max_row + 1):
            metric = kpi_sheet[f"A{row}"].value

            if metric in [
                "Total Revenue",
                "Total Profit",
                "Average Order Value"
            ]:
                kpi_sheet[f"B{row}"].number_format = '$#,##0.00'

            elif metric == "Profit Margin":
                kpi_sheet[f"B{row}"].number_format = '0.0'

        executive_sheet = writer.book["Executive Summary"]

        for row in range(2, executive_sheet.max_row + 1):
            metric = executive_sheet[f"A{row}"].value

            if metric in [
                "Total Revenue",
                "Total Profit",
                "Average Order Value"
            ]:
                executive_sheet[f"B{row}"].number_format = '$#,##0.00'

            elif metric == "Profit Margin":
               executive_sheet[f"B{row}"].number_format = '0.0'

        filtered_sheet = writer.book["Filtered Data"]

        money_columns = {
            "Unit_Price",
            "Unit_Cost",
            "Revenue",
            "Cost",
            "Profit"
        }

        for column in range(1, filtered_sheet.max_column + 1):
            header = filtered_sheet.cell(
                row=1,
                column=column
            ).value

            if header in money_columns:
                for row in range(2, filtered_sheet.max_row + 1):
                    filtered_sheet.cell(
                        row=row,
                        column=column
                    ).number_format = '$#,##0.00'

            elif header == "Profit_Margin":
                for row in range(2, filtered_sheet.max_row + 1):
                    filtered_sheet.cell(
                        row=row,
                        column=column
                    ).number_format = '0.0'

    output.seek(0)

    return output.getvalue()